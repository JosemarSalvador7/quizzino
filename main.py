from customtkinter import IntVar
from random import shuffle
from PIL import Image
import customtkinter as ctk
from openpyxl import load_workbook
import random
from typing import List, Tuple, Dict, Optional, Callable
from functools import partial
import os
import platform
import shutil
import posixpath

if platform.system() == "Linux":
    if not os.path.exists(os.path.join(f"{posixpath.expanduser('~')}", ".Quizzino")):
        shutil.copytree(
            "Quizzino", f"{os.path.join(f'{posixpath.expanduser("~")}', '.Quizzino')}"
        )
        print(f"{os.path.join(f'{posixpath.expanduser("~")}', '.Quizzino')}")
    else:
        print("A pasta já existe.")


# ========== CONSTANTES DE NÍVEL ==========
LEVELS = {
    1: {"name": "Fácil", "questions": 5, "points_per_correct": 1},
    2: {"name": "Médio", "questions": 7, "points_per_correct": 2},
    3: {"name": "Difícil", "questions": 10, "points_per_correct": 3},
}

# ========== FUNÇÕES PURAS E FUNCIONAIS ==========


def load_questions(filepath: str) -> List[List]:
    """Carrega perguntas do Excel de forma funcional"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    # Processamento funcional com list comprehension
    return [
        list(row)
        for row in ws.iter_rows(min_row=2, values_only=True)  # type: ignore
        if any(cell is not None for cell in row)
    ]


def sample_questions(questions: List[List], sample_size: int = 100) -> List[List]:
    """Amostra perguntas de forma funcional"""
    return random.choices(questions, k=sample_size)


def create_color_palette() -> Dict[str, str]:
    """Define a paleta de cores de forma imutável"""
    return {
        "primary": "#0e49c9",
        "secondary": "#042a64",
        "accent": "#464db4",
        "light_accent": "#6495ed",
        "white": "#fffafa",
        "dark_accent": "#363b85",
        "success": "#69e901",
        "error": "#ff0000",
        "gray": "#A0A0A0",
        "dark_gray": "#444343",
        "black": "#000000",
        "level_easy": "#4CAF50",
        "level_medium": "#FF9800",
        "level_hard": "#F44336",
        "level_inactive": "#CCCCCC",
    }


def calculate_score(
    current_score: int, is_correct: bool, level: int, points_per_correct: int
) -> int:
    """Calcula novo score de forma pura considerando nível"""
    if is_correct:
        return current_score + points_per_correct
    return current_score


def get_result_text(score: int, total_questions: int, level: int) -> str:
    """Gera texto de resultado de forma pura"""
    level_name = LEVELS[level]["name"]
    max_points = total_questions * LEVELS[level]["points_per_correct"]

    return (
        f"Nível: {level_name}\n"
        f"Pontuação: {score}/{max_points}\n"
        f"Acertou: {score // LEVELS[level]['points_per_correct']}\n"
        f"Errou: {total_questions - (score // LEVELS[level]['points_per_correct'])}\n"
        f"\nDivirta-se mais!\n"
        f"As perguntas mudam sempre\n"
        f"Venha vencer!"
    )


# ========== CLASSES PARA GERENCIAMENTO DE ESTADO ==========


class QuizState:
    """Gerencia o estado do quiz de forma imutável"""

    def __init__(
        self,
        questions: List[List],
        level: int = 1,
        current_question: int = 0,
        score: int = 0,
    ):
        self.questions = questions
        self.level = level
        self.current_question = current_question
        self.score = score
        self.total_questions = min(LEVELS[level]["questions"], len(questions))
        self.points_per_correct = LEVELS[level]["points_per_correct"]

    def next_question(self, is_correct: bool) -> "QuizState":
        """Retorna novo estado com próxima pergunta"""
        new_score = calculate_score(
            self.score, is_correct, self.level, self.points_per_correct
        )
        return QuizState(
            questions=self.questions,
            level=self.level,
            current_question=self.current_question + 1,
            score=new_score,
        )

    def reset(self, new_level: Optional[int] = None) -> "QuizState":
        """Retorna novo estado resetado"""
        level = new_level if new_level is not None else self.level
        shuffled_questions = self.questions.copy()
        shuffle(shuffled_questions)
        return QuizState(
            questions=shuffled_questions, level=level, current_question=0, score=0
        )

    def change_level(self, new_level: int) -> "QuizState":
        """Muda o nível do quiz"""
        return self.reset(new_level)

    def get_current_question(self) -> Optional[Tuple]:
        """Obtém pergunta atual de forma segura"""
        if self.current_question < self.total_questions:
            question_data = self.questions[self.current_question]
            if len(question_data) >= 6:
                return tuple(question_data[:6])
        return None

    def is_finished(self) -> bool:
        """Verifica se quiz terminou"""
        return self.current_question >= self.total_questions

    def get_level_info(self) -> Dict:
        """Retorna informações do nível atual"""
        return LEVELS[self.level]


# ========== FÁBRICAS DE COMPONENTES ==========


def create_image(filepath: str, size: Tuple[int, int]) -> ctk.CTkImage:
    """Fábrica de imagens"""
    return ctk.CTkImage(
        light_image=Image.open(filepath), dark_image=Image.open(filepath), size=size
    )


def create_button(parent, text: str, command, **kwargs) -> ctk.CTkButton:
    """Fábrica de botões com configuração padrão"""
    defaults = {
        "corner_radius": 10,
        "width": 400,
        "height": 44,
        "font": ("Verdana", 16, "bold"),
    }
    defaults.update(kwargs)
    return ctk.CTkButton(parent, text=text, command=command, **defaults)


def create_label(parent, text: str = "", **kwargs) -> ctk.CTkLabel:
    """Fábrica de labels"""
    return ctk.CTkLabel(parent, text=text, **kwargs)


def create_frame(parent, **kwargs) -> ctk.CTkFrame:
    """Fábrica de frames"""
    return ctk.CTkFrame(parent, **kwargs)


def create_modal(
    parent, text: str, bg_color: str, duration: int, title: str = ""
) -> ctk.CTkToplevel:
    """Fábrica de modais"""
    modal = ctk.CTkToplevel(parent, width=300, height=200, fg_color=bg_color)
    modal.resizable(width=False, height=False)
    if title:
        modal.title(title)

    label = create_label(
        modal,
        text=text,
        font=("Arial", 25, "bold"),
        compound="center",
        anchor="center",
        fg_color="transparent",
    )
    label.pack(pady=50, padx=50)

    modal.after(duration, modal.destroy)
    return modal


# ========== TELA INICIAL ==========


class SplashScreen:
    """Gerencia a tela inicial de apresentação"""

    def __init__(
        self, parent, colors: Dict, fonts: Dict, images: Dict, on_start: Callable
    ):
        self.parent = parent
        self.colors = colors
        self.fonts = fonts
        self.images = images
        self.on_start = on_start

        # Frame 1 - Header
        self.splash1 = create_frame(parent)
        self.splash1.pack(fill="both", pady=2)

        self.splash_title = create_label(
            self.splash1,
            text="Quizino Venha divertir-se",
            width=40,
            height=28,
            fg_color="transparent",
            font=self.fonts["title"],
        )
        self.splash_title.pack()

        # Frame 2 - Conteúdo principal
        self.splash2 = create_frame(parent)
        self.splash2.pack(fill="both", expand=True, ipady=120)

        self.splash_main_text = create_label(
            self.splash2, image=self.images["logo_large"], anchor="center", text=""
        )
        self.splash_main_text.pack(pady=50)

        self.splash_btn = create_button(
            self.splash2,
            text="Jogar Agora",
            font=self.fonts["cursive"],
            command=self.start_game,
            fg_color=self.colors["accent"],
            width=200,
            height=50,
        )
        self.splash_btn.pack()

        # Frame 3 - Footer
        self.splash3 = create_frame(parent)
        self.splash3.pack(fill="both")

        self.splash_footer = create_label(
            self.splash3,
            text="©JEEL Softwares(2025)\nCódigo e Design: João Salvador Paulo",
            width=40,
            height=28,
            fg_color="transparent",
        )
        self.splash_footer.pack()

    def start_game(self):
        """Inicia o jogo e remove a tela inicial"""
        self.splash_btn.pack_forget()
        self.splash1.pack_forget()
        self.splash2.pack_forget()
        self.splash3.pack_forget()

        # Criar footer do jogo
        footer_game = create_frame(self.parent, width=200, height=30)
        footer_game.place(x=153, y=620)

        footer_label = create_label(
            footer_game,
            text="©JEEL Softwares(2025)\nCódigo e Design: João Salvador Paulo",
            font=("Arial", 10),
        )
        footer_label.pack()

        # Chamar callback para iniciar jogo
        self.on_start()


# ========== APLICAÇÃO PRINCIPAL ==========


class QuizinoApp:
    def __init__(self):
        self.app = ctk.CTk()
        self.setup_window()

        # Carregar dados
        self.colors = create_color_palette()
        try:
            questions_data = load_questions(os.path.join("Quizzino/", "questions.xlsx"))
        except Exception as e:
            print(e)
            questions_data = load_questions(
                os.path.join(".Quizzino/", "questions.xlsx")
            )
        sampled_questions = sample_questions(questions_data, 100)

        # Estado inicial (nível 1 - Fácil)
        self.state = QuizState(sampled_questions, level=1)
        try:
            # Assets
            self.images = {
                "logo_large": create_image(
                    os.path.join("Quizzino/", "logo.png"), (200, 200)
                ),
                "logo_medium": create_image(
                    os.path.join("Quizzino/", "logo.png"), (180, 180)
                ),
                "menu": create_image(os.path.join("Quizzino/", "NN.png"), (25, 25)),
            }
        except Exception as e:
            # Assets
            self.images = {
                "logo_large": create_image(
                    os.path.join(".Quizzino/", "logo.png"), (200, 200)
                ),
                "logo_medium": create_image(
                    os.path.join(".Quizzino/", "logo.png"), (180, 180)
                ),
                "menu": create_image(os.path.join(".Quizzino/", "NN.png"), (25, 25)),
            }
            print(e)

        self.fonts = {
            "title": ctk.CTkFont(family="Arial", size=20, weight="bold"),
            "button": ctk.CTkFont(family="Verdana", weight="bold", size=16),
            "cursive": ctk.CTkFont(family="Cursive", size=30, weight="bold"),
            "level": ctk.CTkFont(family="Verdana", size=12, weight="bold"),
        }

        # Controle de visibilidade dos botões
        self.answer_buttons_visible = True

        # Primeiro mostrar tela inicial
        self.show_splash_screen()

    def setup_window(self):
        """Configura a janela principal"""
        self.app.geometry("450x650")
        self.app.resizable(width=False, height=False)
        self.app.title("QUIZINO")

    def show_splash_screen(self):
        """Mostra a tela inicial de apresentação"""
        self.splash = SplashScreen(
            parent=self.app,
            colors=self.colors,
            fonts=self.fonts,
            images=self.images,
            on_start=self.setup_game_ui,
        )

    def setup_game_ui(self):
        """Configura interface do jogo após tela inicial"""
        # Header do jogo
        self.header = create_frame(self.app, height=30)
        self.header.pack(fill="both")

        self.score_label = create_label(
            self.header, text=self.get_score_text(), font=self.fonts["button"]
        )
        self.score_label.place(x=1, y=0)

        # Botão menu
        menu_btn = ctk.CTkButton(
            self.header,
            text="",
            hover_color=self.colors["gray"],
            fg_color="transparent",
            image=self.images["menu"],
            command=self.show_menu,
        )
        menu_btn.place(x=350, y=0)

        # Área de jogo
        self.logo_label = create_label(
            self.app, text="", image=self.images["logo_medium"]
        )
        self.logo_label.pack()

        # Label do nível atual
        self.level_label = create_label(
            self.app,
            text=f"Nível: {self.state.get_level_info()['name']}",
            width=200,
            height=30,
            fg_color=self.get_level_color(),
            font=self.fonts["level"],
            corner_radius=10,
            text_color=self.colors["white"],
        )
        self.level_label.pack(pady=5)

        # Controles de nível
        self.create_level_controls()

        # Label da pergunta
        self.question_label = create_label(
            self.app,
            text="",
            width=400,
            height=62,
            fg_color=self.colors["gray"],
            wraplength=380,
            font=self.fonts["title"],
            corner_radius=10,
            text_color=self.colors["black"],
        )
        self.question_label.pack(pady=10)

        # Frame para botões de resposta (facilita ocultar/revelar)
        self.answers_frame = create_frame(self.app)
        self.answers_frame.pack()

        # Botões de resposta
        self.answer_buttons = []
        for i in range(4):
            btn = create_button(
                self.answers_frame,
                text="",
                command=partial(self.check_answer, i + 1),
                fg_color=self.colors["accent"],
                hover_color=self.colors["dark_accent"],
                state="normal",
            )
            btn.pack(pady=8.5)
            self.answer_buttons.append(btn)

        # Frame para botão jogar novamente
        self.restart_frame = create_frame(self.app)
        # Inicialmente não mostramos este frame

        # Botão jogar novamente
        self.play_again_btn = create_button(
            self.restart_frame,
            text="Jogar Novamente",
            command=self.restart_quiz,
            fg_color=self.colors["accent"],
            hover_color=self.colors["dark_accent"],
        )
        self.play_again_btn.pack()

        # Variável para resposta correta
        self.correct_answer = IntVar()

        # Iniciar o quiz
        self.start_quiz()

    def create_level_controls(self):
        """Cria controles para seleção de nível"""
        level_frame = create_frame(self.app, height=40)
        level_frame.pack(pady=5)

        self.level_buttons = []
        for level_num, level_info in LEVELS.items():
            btn = ctk.CTkButton(
                level_frame,
                text=level_info["name"],
                width=100,
                height=30,
                font=self.fonts["level"],
                fg_color=self.get_level_color(level_num)
                if level_num == self.state.level
                else self.colors["level_inactive"],
                hover_color=self.get_level_color(level_num, dark=True),
                command=partial(self.change_level, level_num),
            )
            btn.pack(side="left", padx=5)
            self.level_buttons.append(btn)

    def get_level_color(self, level: Optional[int] = None, dark: bool = False) -> str:
        """Retorna a cor do nível"""
        if level is None:
            level = self.state.level

        color_map = {
            1: self.colors["level_easy"],
            2: self.colors["level_medium"],
            3: self.colors["level_hard"],
        }

        base_color = color_map.get(level, self.colors["accent"])

        if dark:
            # Escurece a cor para hover
            return self.darken_color(base_color)

        return base_color

    def darken_color(self, color: str) -> str:
        """Escurece uma cor hex"""
        # Implementação simplificada para escurecer cores
        return color  # Em uma implementação real, converteríamos e escureceríamos

    def get_score_text(self) -> str:
        """Obtém texto do score de forma pura"""
        level_info = self.state.get_level_info()
        max_points = self.state.total_questions * level_info["points_per_correct"]
        return f"Pontuação: {self.state.score}/{max_points}"

    def show_menu(self):
        """Mostra menu sobre"""
        modal = ctk.CTkToplevel(self.app, width=200, height=200)
        modal.title("Sobre")

        tabview = ctk.CTkTabview(
            modal, segmented_button_selected_color=self.colors["light_accent"]
        )
        tabview.pack(padx=10, pady=10)

        # Abas
        tabs = ["Sobre", "Autores", "Níveis"]
        for tab_name in tabs:
            tabview.add(tab_name)

        tabview.set("Sobre")

        # Conteúdo das abas - Sobre
        about_content = [
            ("Quizino - Versão 2.0", {"font": ("Verdana", 16, "bold")}),
            (
                "Quizino \nVenha aprender e divertir-se\nQuizino é um projeto JEEL Softwares",
                {"font": ("Verdana", 14, "bold")},
            ),
            ("\nSistema de Níveis:", {"font": ("Verdana", 12, "bold")}),
            ("• Fácil: 5 questões, 1 ponto cada", {"font": ("Verdana", 11)}),
            ("• Médio: 7 questões, 2 pontos cada", {"font": ("Verdana", 11)}),
            ("• Difícil: 10 questões, 3 pontos cada", {"font": ("Verdana", 11)}),
        ]

        for text, kwargs in about_content:
            label = create_label(tabview.tab("Sobre"), text=text, **kwargs)
            label.pack()

        # CORREÇÃO: Remover fg_color='transparent' da chamada da função
        # Conteúdo das abas - Autores
        authors_content = [
            ("DESENVOLVIDO POR", {"font": ("Verdana", 16, "bold")}),
            (
                "João Salvador Paulo",
                {
                    "text_color": self.colors["white"],
                    "fg_color": self.colors["dark_gray"],
                    "corner_radius": 10,
                    "font": ("Verdana", 16, "bold"),
                },
            ),
            ("\nCódigo e Design", {"font": ("Verdana", 12, "bold")}),
            (
                "João Salvador Paulo",
                {
                    "text_color": self.colors["white"],
                    "fg_color": self.colors["dark_gray"],
                    "corner_radius": 10,
                    "font": ("Verdana", 12),
                },
            ),
        ]

        for text, kwargs in authors_content:
            label = create_label(tabview.tab("Autores"), text=text, **kwargs)
            label.pack()

        # CORREÇÃO: Remover fg_color='transparent' da chamada da função
        # Conteúdo da aba Níveis
        levels_content = [
            ("SISTEMA DE NÍVEIS", {"font": ("Verdana", 16, "bold")}),
            (
                "\nFácil (Verde):",
                {
                    "font": ("Verdana", 12, "bold"),
                    "text_color": self.colors["level_easy"],
                },
            ),
            ("• 5 perguntas", {"font": ("Verdana", 11)}),
            ("• 1 ponto por acerto", {"font": ("Verdana", 11)}),
            (
                "\nMédio (Laranja):",
                {
                    "font": ("Verdana", 12, "bold"),
                    "text_color": self.colors["level_medium"],
                },
            ),
            ("• 7 perguntas", {"font": ("Verdana", 11)}),
            ("• 2 pontos por acerto", {"font": ("Verdana", 11)}),
            (
                "\nDifícil (Vermelho):",
                {
                    "font": ("Verdana", 12, "bold"),
                    "text_color": self.colors["level_hard"],
                },
            ),
            ("• 10 perguntas", {"font": ("Verdana", 11)}),
            ("• 3 pontos por acerto", {"font": ("Verdana", 11)}),
        ]

        for text, kwargs in levels_content:
            label = create_label(tabview.tab("Níveis"), text=text, **kwargs)
            label.pack()

    def change_level(self, new_level: int):
        """Altera o nível atual"""
        if self.state.level != new_level:
            # Muda para novo nível
            self.state = self.state.change_level(new_level)

            # Atualiza cores dos botões de nível
            for i, btn in enumerate(self.level_buttons):
                level_num = i + 1
                btn.configure(
                    fg_color=self.get_level_color(level_num)
                    if level_num == new_level
                    else self.colors["level_inactive"]
                )

            # Atualiza label do nível
            self.level_label.configure(
                text=f"Nível: {self.state.get_level_info()['name']}",
                fg_color=self.get_level_color(),
            )

            # Reinicia o quiz no novo nível
            self.restart_quiz()

    def check_answer(self, answer: int):
        """Verifica resposta"""
        # Captura a pergunta atual antes de avançar o estado
        current_q = self.state.get_current_question()
        is_correct = answer == self.correct_answer.get()

        # Atualizar estado
        self.state = self.state.next_question(is_correct)

        # Mostrar feedback
        if is_correct:
            self.show_success_modal()
        else:
            if current_q:
                correct_idx = self.correct_answer.get() - 1
                correct_text = current_q[correct_idx + 1]  # +1 para pular pergunta
                self.show_error_modal(self.correct_answer.get(), correct_text)

        # Atualizar UI
        self.update_ui()

        # Verificar se terminou
        if self.state.is_finished():
            self.show_final_results()

    def update_ui(self):
        """Atualiza interface do usuário"""
        self.score_label.configure(text=self.get_score_text())

        if not self.state.is_finished():
            self.display_question()

    def display_question(self):
        """Exibe pergunta atual"""
        question_data = self.state.get_current_question()
        if not question_data:
            return

        question, *options, answer = question_data

        # Atualizar label da pergunta
        self.question_label.configure(text=question)

        # Atualizar botões
        for i, (btn, option) in enumerate(zip(self.answer_buttons, options)):
            btn.configure(text=option, state="normal")

        # Definir resposta correta
        self.correct_answer.set(answer)

    def show_success_modal(self):
        """Mostra modal de acerto"""
        points = self.state.points_per_correct
        create_modal(
            self.app,
            f"Acertou!\n+{points} ponto{'s' if points > 1 else ''}",
            self.colors["success"],
            1000,
            "Acertou!",
        )

    def show_error_modal(self, position: int, correct_answer: str):
        """Mostra modal de erro"""
        text = f"Errou!\nResposta correta:\nOpção {position}: {correct_answer}"
        create_modal(self.app, text, self.colors["error"], 3500, "ERROU")

    def show_final_results(self):
        """Mostra resultados finais"""
        text = get_result_text(
            self.state.score, self.state.total_questions, self.state.level
        )
        create_modal(self.app, text, self.get_level_color(), 6500, "Resultado Final")

        # PASSO 1: Limpar a tela atual
        self.hide_game_elements()

        # PASSO 2: Mostrar botão "Jogar Novamente"
        self.show_restart_screen()

    def hide_game_elements(self):
        """Esconde elementos do jogo quando termina"""
        if self.answer_buttons_visible:
            # Esconde botões de resposta
            self.answers_frame.pack_forget()

            # Limpa texto da pergunta
            self.question_label.configure(text="")

            # Atualiza flag
            self.answer_buttons_visible = False

    def show_restart_screen(self):
        """Mostra tela de reinício com botão Jogar Novamente"""
        # Remove elementos do jogo se ainda estiverem visíveis
        self.hide_game_elements()

        # Mostra frame do botão jogar novamente
        self.restart_frame.pack(pady=20)

        # Atualiza score label para mostrar resultado final
        level_info = self.state.get_level_info()
        max_points = self.state.total_questions * level_info["points_per_correct"]
        self.score_label.configure(text=f"FINAL: {self.state.score}/{max_points}")

    def show_game_elements(self):
        """Mostra elementos do jogo novamente"""
        if not self.answer_buttons_visible:
            # Mostra frame dos botões de resposta
            self.answers_frame.pack()

            # Atualiza flag
            self.answer_buttons_visible = True

    def restart_quiz(self):
        """Reinicia o quiz"""
        # PASSO 1: Esconder botão "Jogar Novamente"
        self.restart_frame.pack_forget()

        # PASSO 2: Resetar estado
        self.state = self.state.reset()

        # PASSO 3: Mostrar elementos do jogo novamente
        self.show_game_elements()

        # PASSO 4: Atualizar UI
        self.score_label.configure(text=self.get_score_text())
        self.level_label.configure(
            text=f"Nível: {self.state.get_level_info()['name']}",
            fg_color=self.get_level_color(),
        )

        # PASSO 5: Começar novo quiz
        self.start_quiz()

    def start_quiz(self):
        """Inicia o quiz"""
        self.display_question()

    def run(self):
        """Executa aplicação"""
        self.app.mainloop()


# ========== EXECUÇÃO PRINCIPAL ==========


def main():
    """Função principal"""
    app = QuizinoApp()
    app.run()


if __name__ == "__main__":
    main()
